import smtplib
import urllib.request as urllib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
from bs4 import BeautifulSoup
from openpyxl import workbook, worksheet, load_workbook
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mtick
from matplotlib.figure import figaspect
from openpyxl.utils import get_column_letter
import numpy as np

filePath = 'Exchange Rate Fluctuation COP.xlsx'
img1Path = 'image004.jpg'
img2Path = 'image003.jpg'


def send_email(fromAdr, toAdr, subject, ws):
    msg = MIMEMultipart()
    msg['From'] = fromAdr
    msg['To'] = toAdr
    msg['Subject'] = subject

    body = bodyTable(ws)
    body += '<hr><img src="cid:%s"> <hr> <img src="cid:%s">' % (img2Path, img1Path)
    msg.attach(MIMEText(body, 'html '))

    fp = open(img1Path, 'rb')
    img = MIMEImage(fp.read())
    fp.close()
    img.add_header('Content-ID', '<{}>'.format(img1Path))
    msg.attach(img)

    fp = open(img2Path, 'rb')
    img = MIMEImage(fp.read())
    fp.close()
    img.add_header('Content-ID', '<{}>'.format(img2Path))
    msg.attach(img)

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(filePath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % filePath)
    msg.attach(part)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(fromAdr, 'samaria95')

    text = msg.as_string()
    server.sendmail(fromAdr, toAdr, text)
    server.quit()


def bodyTable(ws: worksheet.Worksheet):
    bold = True
    body = ''
    body += '<table border="solid thin">'

    body += '<tr>'
    body += '<td colspan ="4"><b>Today</b></td>'
    body += '<td colspan ="4"><b>Monthly Avg</b></td>'
    body += '</tr>'

    for j in range(2, 4):
        body += '<tr>'
        for i in range(10, 18):
            col = get_column_letter(i)
            body += '<td>'
            if bold:
                body += '<b>'
            td = ws[col + str(j)].value
            if isinstance(td, float):
                if td < 1:
                    body += ('{0:.1f}%'.format(td * 100))
                else:
                    body += "{0:,.0f}".format(td)
            elif isinstance(td, datetime):
                body += td.strftime('%d/%m/%Y')
            else:
                body += str(td)
            if bold:
                body += '</b>'
            body += '</td>'
        if bold:
            body += '<td>' + '<b>Budget FY19</b>' + '</td>'
        else:
            body += '<td>' + '{0:,.0f}'.format(ws['R2'].value) + '</td>'

        body += '</tr>'
        bold = False

    body += '</table>'
    return body


def getTRM():
    page = urllib.urlopen('https://www.superfinanciera.gov.co/jsp/index.jsf').read()
    soup = BeautifulSoup(page, 'lxml')
    s1 = soup.find("td", text="TRM").parent
    s2 = s1.find_all("td")[1]
    tasaCambioString = s2.text
    for char in tasaCambioString:
        if char in ' ,$':
            tasaCambioString = tasaCambioString.replace(char, '')
    tasaCambio = float(tasaCambioString)
    if tasaCambio < 1500 or tasaCambio > 4000:
        raise Exception('Valor fuera de rango')
    return tasaCambio


def encontrarFila(ws: worksheet.Worksheet) -> int:
    date = datetime.now()
    date = date.date()

    for i in range(100, 2000):
        cell = ws['e' + str(i)]
        if cell.value is not None and cell.value.date() == date:
            return i


def actualizarPromMes(ws: worksheet.Worksheet, fila):
    promMes2019 = encontrarPromedioMesActual(ws, fila, 'E', 'F')
    promMes2018 = encontrarPromedioMesActual(ws, fila, 'C', 'D')
    ws['N3'] = promMes2018
    ws['O3'] = promMes2019


def encontrarPromedioMesActual(ws: worksheet.Worksheet, fila, columnaMes, columnaTRM):
    filaI = fila
    filaF = fila
    mes = ws[columnaMes + str(fila)].value.month

    suma = ws[columnaTRM + str(fila)].value
    count = 1
    for i in range(32):
        filaI = filaI - 1
        mes1 = ws[columnaMes + str(filaI)].value.month

        if mes1 == mes and ws[columnaTRM + str(filaI)].value is not None:
            suma += ws[columnaTRM + str(filaI)].value
            count += 1
        elif mes1 != mes:
            break

    for i in range(32):
        filaF = filaF + 1
        mes2 = ws[columnaMes + str(filaF)].value.month
        if mes2 == mes and ws[columnaTRM + str(filaF)].value is not None:
            suma += ws[columnaTRM + str(filaF)].value
            count += 1
        if mes2 != mes:
            break

    return suma / count


def calcularDifyVarDeFila(ws: worksheet.Worksheet, fila):
    trm2019 = ws['F' + str(fila)].value
    trm2018 = ws['D' + str(fila)].value
    if trm2019 is not None and trm2018 is not None:
        ws['G' + str(fila)] = trm2019 - trm2018
        ws['H' + str(fila)] = trm2019 / trm2018 - 1


def calcularDif(ws: worksheet.Worksheet):
    for i in range(737, 900):
        calcularDifyVarDeFila(ws, i)


def graficar(ws: worksheet.Worksheet):
    target = ws['R2'].value

    dates = []
    e2017 = []
    e2018 = []
    e2019 = []

    for i in range(737, 1102):
        date = ws['A' + str(i)].value
        valor2017 = ws['B' + str(i)].value
        valor2018 = ws['D' + str(i)].value
        valor2019 = ws['F' + str(i)].value

        dates.append(date)
        e2017.append(valor2017)
        e2018.append(valor2018)
        e2019.append(valor2019)

    w, h = figaspect(1 / 2)
    fig, ax1 = plt.subplots(figsize=(w, h))
    ax1.plot(dates, e2017, label='FY2017')
    ax1.plot(dates, e2018, label='FY2018')
    ax1.plot(dates, e2019, label='FY2019')

    ax1.axhline(y=target, color='r', label='Budget FY2019')

    ax1.xaxis.set_major_locator(mdates.MonthLocator())
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%d-%b"))

    plt.xticks(rotation=90)
    ax1.xaxis.set_minor_locator(mdates.MonthLocator())

    plt.tick_params(axis='both', which='major', labelsize=8)
    plt.tick_params(axis='both', which='minor', labelsize=8)

    fmt = '${x:,.0f}'
    tick = mtick.StrMethodFormatter(fmt)

    minY = min(x for x in (e2017 + e2018 + e2019) if x is not None) // 500 * 500
    maxY = (max(x for x in (e2017 + e2018 + e2019) if x is not None) // 500 + 1) * 500

    ax1.yaxis.set_ticks(np.arange(minY, maxY, 50))
    ax1.yaxis.set_major_formatter(tick)
    plt.grid()

    leg = plt.legend(bbox_to_anchor=(1, -0.2), ncol=4)

    ax1.set_title('Exchange Rate Fluctuation')
    fig.savefig('image003.jpg', bbox_extra_artists=(leg,), bbox_inches='tight')


if __name__ == '__main__':
    tasaCambio = getTRM()
    wb = load_workbook(filename=filePath)
    ws = wb['Data']
    fila = encontrarFila(ws)

    ws['F' + str(fila)] = tasaCambio

    calcularDif(ws)

    ws['J3'] = ws['D' + str(fila)].value
    ws['K3'] = ws['F' + str(fila)].value
    ws['J2'] = ws['C' + str(fila)].value
    ws['K2'] = ws['E' + str(fila)].value

    actualizarPromMes(ws, fila)
    ws['L3'] = ws['K3'].value / ws['J3'].value - 1
    ws['M3'] = ws['K3'].value / ws['R2'].value - 1
    ws['P3'] = ws['O3'].value / ws['N3'].value - 1
    ws['Q3'] = ws['O3'].value / ws['R2'].value - 1

    wb.save(filePath)
    wb = load_workbook(filename=filePath)
    graficar(ws)

    send_email('dgiraldom95@gmail.com', 'lgallo@pricesmart.com', 'Exchange Rate Test', ws)
    # send_email('dgiraldom95@gmail.com', 'dgiraldom95@gmail.com', 'Exchange Rate Test', ws)
